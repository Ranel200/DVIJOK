"""XLSX preview/import: validation, atomicity and tenant isolation."""

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.core.config import settings
from app.modules.organizations.models import Organization
from app.modules.services.models import Service
from app.shared.enums import LegalForm, OrganizationStatus, ServiceCategory, TaxSystem
from tests.conftest import API

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx(rows, headers=("Услуга", "Стоимость")) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def test_download_template_has_required_headers_and_numeric_price(auth_client):
    response = await auth_client.get(f"{API}/services/import/template")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_TYPE)

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    sheet = workbook.active
    assert [sheet["A1"].value, sheet["B1"].value] == ["Услуга", "Стоимость"]
    assert isinstance(sheet["B2"].value, (int, float))
    assert sheet.freeze_panes == "A2"


async def test_preview_is_read_only_and_apply_uses_safe_defaults(auth_client):
    content = _xlsx([("Замена масла", 2500), ("Диагностика", "1 500,50")])
    preview = await auth_client.post(
        f"{API}/services/import/preview",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["valid"] is True
    assert preview.json()["valid_rows"] == 2
    assert preview.json()["imported_rows"] == 0

    before = await auth_client.get(f"{API}/services", params={"query": "Замена масла"})
    assert before.json()["total"] == 0

    applied = await auth_client.post(
        f"{API}/services/import",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["imported_rows"] == 2

    listing = await auth_client.get(f"{API}/services", params={"query": "Замена масла"})
    imported = listing.json()["items"][0]
    assert imported["category"] == "other"
    assert imported["duration_minutes"] == 60
    assert imported["labor_hours"] == "0.00"
    assert imported["base_price"] == "2500.00"

    duplicate = await auth_client.post(
        f"{API}/services/import",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert duplicate.status_code == 422
    assert duplicate.json()["imported_rows"] == 0
    assert "уже существует" in duplicate.json()["errors"][0]["message"]


async def test_invalid_rows_report_excel_numbers_and_import_is_atomic(auth_client):
    content = _xlsx(
        [
            ("Шиномонтаж", 3000),
            ("  шиномонтаж  ", 3500),
            ("", 100),
            ("Плохая цена", -1),
        ]
    )
    response = await auth_client.post(
        f"{API}/services/import",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert response.status_code == 422, response.text
    body = response.json()
    assert body["valid"] is False
    assert body["imported_rows"] == 0
    assert {error["row_number"] for error in body["errors"]} == {3, 4, 5}

    listing = await auth_client.get(f"{API}/services", params={"query": "Шиномонтаж"})
    assert listing.json()["total"] == 0


async def test_wrong_headers_and_formula_price_are_rejected(auth_client):
    bad_headers = _xlsx([("Услуга", 100)], headers=("Название", "Цена"))
    header_response = await auth_client.post(
        f"{API}/services/import/preview",
        files={"file": ("services.xlsx", bad_headers, XLSX_TYPE)},
    )
    assert header_response.json()["valid"] is False
    assert header_response.json()["errors"][0]["row_number"] == 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Услуга", "Стоимость"])
    sheet.append(["Формульная услуга", "=100+200"])
    output = BytesIO()
    workbook.save(output)
    formula_response = await auth_client.post(
        f"{API}/services/import/preview",
        files={"file": ("services.xlsx", output.getvalue(), XLSX_TYPE)},
    )
    assert formula_response.json()["valid"] is False
    assert formula_response.json()["errors"][0]["row_number"] == 2


async def test_same_name_in_another_tenant_does_not_block_import(
    auth_client, session_factory
):
    async with session_factory() as session:
        other = Organization(
            name="Другой сервис",
            inn="6677889900",
            tax_system=TaxSystem.USN,
            legal_form=LegalForm.OOO,
            legal_address="Москва",
            phone="+70001112233",
            status=OrganizationStatus.ACTIVE,
        )
        session.add(other)
        await session.flush()
        session.add(
            Service(
                organization_id=other.id,
                name="Полировка",
                category=ServiceCategory.OTHER,
                base_price=5000,
                labor_hours=0,
                duration_minutes=60,
            )
        )
        await session.commit()

    content = _xlsx([("Полировка", 4500)])
    response = await auth_client.post(
        f"{API}/services/import",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert response.status_code == 200, response.text
    assert response.json()["imported_rows"] == 1

    async with session_factory() as session:
        result = await session.execute(
            select(Service).where(Service.name == "Полировка")
        )
        assert len(result.scalars().all()) == 2


async def test_file_size_limit_is_reported(auth_client, monkeypatch):
    content = _xlsx([("Большой файл", 100)])
    monkeypatch.setattr(settings, "SERVICE_IMPORT_MAX_FILE_BYTES", 100)
    response = await auth_client.post(
        f"{API}/services/import/preview",
        files={"file": ("services.xlsx", content, XLSX_TYPE)},
    )
    assert response.status_code == 200
    assert response.json()["valid"] is False
    assert "Размер файла превышает" in response.json()["errors"][0]["message"]
