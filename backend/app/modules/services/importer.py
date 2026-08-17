"""XLSX template and strict atomic service-catalog import."""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.config import settings
from app.modules.services.schemas import (
    ServiceImportError,
    ServiceImportReport,
    ServiceImportRow,
)

EXPECTED_HEADERS = {"услуга": "name", "стоимость": "base_price"}
MAX_PRICE = Decimal("99999999.99")


def normalize_name(value: str) -> str:
    return " ".join(value.split()).casefold()


def build_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Услуги"
    sheet.append(["Услуга", "Стоимость"])
    sheet.append(["Замена масла", 2500])
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 18
    sheet["B2"].number_format = '#,##0.00" ₽"'
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    table = Table(displayName="ServicesImport", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@dataclass
class ParsedImport:
    report: ServiceImportReport


def _parse_price(value: object) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError("Стоимость обязательна")
    raw = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        price = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Стоимость должна быть числом") from exc
    if not price.is_finite() or price <= 0:
        raise ValueError("Стоимость должна быть положительным числом")
    if price > MAX_PRICE:
        raise ValueError(f"Стоимость не должна превышать {MAX_PRICE}")
    exponent = price.as_tuple().exponent
    if isinstance(exponent, int) and exponent < -2:
        raise ValueError("Стоимость должна содержать не более двух знаков после запятой")
    return price.quantize(Decimal("0.01"))


def parse_workbook(data: bytes, existing_names: set[str]) -> ParsedImport:
    errors: list[ServiceImportError] = []
    rows: list[ServiceImportRow] = []
    try:
        with ZipFile(BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > 1000 or sum(item.file_size for item in entries) > 20 * 1024 * 1024:
                raise ValueError("XLSX содержит слишком много распакованных данных")
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except (BadZipFile, OSError, ValueError, KeyError):
        return ParsedImport(
            ServiceImportReport(
                valid=False,
                total_rows=0,
                valid_rows=0,
                errors=[
                    ServiceImportError(
                        field="file",
                        message=(
                            "Файл повреждён, слишком велик после распаковки или не является XLSX"
                        ),
                    )
                ],
                rows=[],
            )
        )

    sheet = workbook.active
    header_values = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1), ())
    ]
    header_map: dict[str, int] = {}
    for index, header in enumerate(header_values):
        normalized = header.casefold()
        if normalized:
            if normalized in header_map:
                errors.append(
                    ServiceImportError(
                        row_number=1,
                        field="headers",
                        message=f"Заголовок «{header}» указан повторно",
                    )
                )
            header_map[normalized] = index
    missing = set(EXPECTED_HEADERS) - set(header_map)
    extra = set(header_map) - set(EXPECTED_HEADERS)
    if missing:
        errors.append(
            ServiceImportError(
                row_number=1,
                field="headers",
                message="Отсутствуют столбцы: "
                + ", ".join(f"«{header.title()}»" for header in sorted(missing)),
            )
        )
    if extra:
        errors.append(
            ServiceImportError(
                row_number=1,
                field="headers",
                message="Неизвестные столбцы: "
                + ", ".join(f"«{header}»" for header in sorted(extra)),
            )
        )
    if missing or extra:
        workbook.close()
        return ParsedImport(
            ServiceImportReport(
                valid=False,
                total_rows=0,
                valid_rows=0,
                errors=errors,
                rows=[],
            )
        )

    seen: dict[str, int] = {}
    total_rows = 0
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        total_rows += 1
        if total_rows > settings.SERVICE_IMPORT_MAX_ROWS:
            errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="file",
                    message=f"Допустимо не более {settings.SERVICE_IMPORT_MAX_ROWS} строк",
                )
            )
            break

        name_value = cells[header_map["услуга"]].value
        price_cell = cells[header_map["стоимость"]]
        row_errors: list[ServiceImportError] = []
        name = " ".join(str(name_value or "").split())
        if not name:
            row_errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="Услуга",
                    message="Название услуги обязательно",
                )
            )
        elif len(name) > 255:
            row_errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="Услуга",
                    message="Название не должно превышать 255 символов",
                )
            )
        normalized_name = normalize_name(name)
        if normalized_name and normalized_name in seen:
            row_errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="Услуга",
                    message=f"Дубликат строки {seen[normalized_name]}",
                )
            )
        elif normalized_name and normalized_name in existing_names:
            row_errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="Услуга",
                    message="Услуга с таким названием уже существует",
                )
            )
        if normalized_name:
            seen.setdefault(normalized_name, row_number)

        if price_cell.data_type == "f":
            row_errors.append(
                ServiceImportError(
                    row_number=row_number,
                    field="Стоимость",
                    message="Формулы не поддерживаются; вставьте вычисленное число",
                )
            )
            price = None
        else:
            try:
                price = _parse_price(price_cell.value)
            except ValueError as exc:
                row_errors.append(
                    ServiceImportError(
                        row_number=row_number,
                        field="Стоимость",
                        message=str(exc),
                    )
                )
                price = None
        errors.extend(row_errors)
        if not row_errors and price is not None:
            rows.append(
                ServiceImportRow(
                    row_number=row_number,
                    name=name,
                    base_price=price,
                )
            )
    workbook.close()
    if total_rows == 0:
        errors.append(
            ServiceImportError(
                field="file",
                message="Файл не содержит строк услуг",
            )
        )
    return ParsedImport(
        ServiceImportReport(
            valid=not errors,
            total_rows=total_rows,
            valid_rows=len(rows),
            errors=errors,
            rows=rows,
        )
    )
